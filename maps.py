#!/usr/bin/env python
import csv
import sys
import os
from numpy import arange
from openpyxl import load_workbook, Workbook
from shutil import move
import datetime
from time import sleep


assaylist = []


class Assay:
    def __init__(self):
        self.disease = " "
        self.geltype = " "
        self.digested = " "
        self.conditions = " "
        self.samples = []
        self.locations = []
        self.paired_assay = ''
    def set_disease(self, disease):
        self.disease = disease
    def set_geltype(self, geltype):
        self.geltype = geltype
    def set_digested(self, digested):
        self.digested = digested
    def set_conditions(self, conditions):
        self.conditions = conditions
    def add_sample(self, sample):
        self.samples.append(sample)
    def add_location(self, location):
        self.locations.append(location)
    def paired(self, assay):
        self.paired_assay = assay
    def set_A_allele(self, size):
        self.A_allele = size
    def set_AB_allele(self, size):
        self.AB_allele = size
    def set_B_allele(self, size):
        self.B_allele = size


def main():
    original_directory = os.getcwd()
    # path to folder for platemap
    path = getpath()
    # gets platemap from folder
    platemap = getplatemap(path)
    # gets specific sheet from platemap
    sheet1 = platemapsheet1(platemap)
    sheet2 = platemapsheet2(platemap)
    # gets all assays, samples
    objects = getassays(sheet1)
    # get assay info
    getinfo(sheet2, objects)
    get_paired_assays(sheet2, objects, original_directory)
    # outputs assays, samplecount, and reagents to template copy
    output(path, objects, original_directory)
    

def getpath():
    # path to folder for platemap
    yesno = "x"
    while yesno.lower() not in "yesno":
        print("Attempt automatic folder detection?")
        yesno = input(":")
        if yesno.lower() not in "yesno":
            print("What?")
    if yesno.lower() in "yes":
        today = datetime.date.today()
        monday_date = today - datetime.timedelta(days=today.weekday())
        week_of = "Week of {}-{}-{}".format(monday_date.strftime("%m"), monday_date.strftime("%d"), monday_date.strftime("%y"))
        if today.weekday() != 0 and today.weekday() != 2:
            previous_date = today - datetime.timedelta(days=1)
        else:
            previous_date = today
        previous_folder = "{}-{}-{}".format(previous_date.strftime("%m"), previous_date.strftime("%d"), previous_date.strftime("%y"))
        try:
            os.chdir('../../Current Year/{}/{}'.format(week_of, previous_folder))
            path = os.getcwd()
        except FileNotFoundError:
            print("Automatic gel map detection failed.")
            path = manual_directory()
    else:
        path = manual_directory()
    return path


def manual_directory():
    while True:
            path = input("Enter the path of your platemap file: ")
            if not os.path.isdir(path):
                print("Not a valid directory")
            else:
                break
    return path


def getplatemap(path):
    # automated detection of most recent directory, can be opted out of
    filelist= []
    for file in os.listdir(path):
        filelist.append(file)
    for file in filelist:
        if 'gel_' in file:
            print('Is this the platemap?')
            print(file)
            yesno = input(':')
            if yesno.lower() in "yes":
                path1 = path + '\\' + file
                return path1
            elif yesno.lower() in 'no':
                while True:
                    print('Here is a list of the files in that directory.')
                    print(filelist)
                    try:
                        filenum = int(input('Which would you like to use? (enter a number corresponding to the order of the files):'))
                        platemap = filelist[filenum - 1]
                        print("You selected '{}'. Are you sure?".format(platemap))
                        yesno = input(':')
                        if yesno.lower() in "yes":
                            break
                        elif yesno.lower() not in 'no':
                            print("this is a yes or no question")
                    except (TypeError, ValueError):
                        print("Use a number dummy")
                    except IndexError:
                        print("Count much? Use a number that refers to one of the files")
                path1 =  path + "\\" + platemap
                return path1        
            else:
                print('You broke it. Start over and answer yes-no questions with yes-no answers')
                sleep(3)
                quit()


def fileselect(filelist):
    # gets user input for which file in the folder is the platemap
    while True: 
        try:
            filenum = int(input('Which would you like to use? (enter a number corresponding to the order of the files):'))
            platemap = filelist[filenum - 1]
            print("You selected '{}'. Are you sure?".format(platemap))
            yesno = input(':')
            if yesno.lower() in "yes":
                return platemap
            elif yesno.lower() not in 'no':
                print("this is a yes or no question")
        except (TypeError, ValueError):
            print("Use a number dummy")
        except IndexError:
            print("Count much? Use a number that refers to one of the files")        


def platemapsheet1(path1):
    tries = 0
    while tries < 3:
        try:
            wb = load_workbook(filename=path1, data_only=True)
            ws = wb.worksheets[0]
        except PermissionError:
            print('Close out of the platemap. The program will try again in 5 seconds.')
            sleep(5)
            tries += 1
        if tries == 2:
            print('Alright, clearly nothing is changing. Make sure the platemap is closed and rerun the program')
            sleep(5)
            quit()
    return ws


def platemapsheet2(path1):
    # gets reaction information sheet
    wb = load_workbook(filename=path1, data_only=True)
    ws = wb.worksheets[1]
    return ws


def getinfo(ws, objects):
    # storing information from reaction sheet
    for i in range(1, ws.max_column+1):
        primerpair = ws.cell(row=4,column=i).value
        if primerpair in objects.keys():
            objects[primerpair].set_disease(ws.cell(row=1,column=i).value)
            objects[primerpair].set_geltype(ws.cell(row=3,column=i).value)
            objects[primerpair].set_A_allele(ws.cell(row=7,column=i).value)
            objects[primerpair].set_AB_allele(ws.cell(row=8,column=i).value)
            objects[primerpair].set_B_allele(ws.cell(row=9,column=i).value)
            if ws.cell(row=10,column=i).value != None:
                objects[primerpair].set_digested(True)
            else:
                objects[primerpair].set_digested(False)
            # A few special cases for LIMS issues
            if primerpair == "ATP7B_112GA_RD_2" or primerpair == "ATP7B_112GA_RD_4" or primerpair == "ATP7B_112GA_RD_5":
                objects[primerpair].set_disease("Copper Toxicosis (Labrador Retriever Type) ATP7B")


def get_paired_assays(ws, objects, original_directory):
    # swaps back to original directory to grab reference sheet for paired assays
    # this is done after changing directories to read platemap so that only assays that are present are searched for
    cwd = os.getcwd()
    os.chdir(original_directory)
    wb = load_workbook(filename='Map Making Key.xlsx', data_only=True)
    ws= wb.worksheets[2]
    max_row = ws.max_row+1
    for a in range(2, max_row):
        if ws.cell(row=a,column=3).value == True and (ws.cell(row=a,column=2).value in assaylist):
            objects[ws.cell(row=a,column=2).value].paired(ws.cell(row=a,column=5).value)
    os.chdir(cwd)


def getassays(ws):
    # gets all assays and initial samplecount
    bottom_rows = list(arange(15,431,16))
    assay_rows = sorted(list(arange(2,418,16)) + bottom_rows)
    rowcounter = 0
    # use a variable column titles to work with assay names before moving to a new object in the event of an assay being present under another assay
    columntitles = []
    # temporary storage for samples tucked under other assays. These are appended to the end of the sample list so sample order is retained, even if these
    # tucked samples appear first on the worklist
    tucked_samples = []
    tucked_locations = []
    # dictionary for assays
    objects = {}
    bottomrowcounter = 0
    plate = 0
    botrowlen = len(bottom_rows)
    for row in ws.values:
        # Resets a column value for each new row. The column value allows me to iterate trough each member of the list I made from the row
        # I start at 4 because I know the values I care about will start in column 4
        columncounter = 3
        rowcounter += 1
        # If the row of the sheet will have assay information according to the predefined assay rows
        if rowcounter in assay_rows:
            columntitles.clear()     
            while True:
                cell = '{}{}'.format(chr(64 + columncounter), rowcounter)
                cellvalue = ws[cell].value
                # Will add the value in the current cell to the assay list if it is not already present in the list
                if cellvalue not in assaylist and cellvalue != None:
                    assaylist.append(cellvalue)
                    if cellvalue not in objects.keys():
                        objects[cellvalue] = Assay()
                columntitles.append(cellvalue)    
                columncounter += 1
                if columncounter == 15:
                    break
        if (rowcounter - 7) % 16 == 0:
            # new plate
            plate += 1
            for a in range(12):
                # each column
                tucked = False
                temprowcounter = rowcounter
                current_column = columntitles[columncounter-3]
                # will be the column title underneath, if present
                previous_column = columntitles[columncounter-4]
                if current_column != previous_column:
                    # meaning we've reached the end of an assay, time to add any tucked samples that were temporarily saved
                    if len(tucked_samples) != 0:
                        for sample in tucked_samples:
                            (objects[previous_column]).add_sample(sample)
                        for location in tucked_locations:
                            (objects[previous_column]).add_location(location)
                        tucked_samples.clear()
                        tucked_locations.clear()
                for b in range(8):
                    # excel cell format
                    cell = '{}{}'.format(chr(64 + columncounter), temprowcounter)
                    # a note that says where the sample is on which plate
                    plateloc = 'P' + str(plate) + ' ' + chr(b + 65) + str(columncounter-2)
                    cellvalue = ws[cell].value
                    current_column = columntitles[columncounter-3]
                    if current_column != None:
                        if cellvalue != ' ' and cellvalue != None:
                            if tucked:
                                # send sample info to temprary storage
                                tucked_samples.append(cellvalue)
                                tucked_locations.append(plateloc)
                            if len(objects[current_column].samples) == 0 and cellvalue[0] != 'R' and len(tucked_samples) == 0:
                                # if the first sample of the new assay doesn't start with 'R', indicating a control, assumed to be a tucked assay
                                tucked_samples.append(cellvalue)
                                tucked_locations.append(plateloc)
                                tucked = True
                            if not tucked:
                                # add directly to assay info
                                (objects[current_column]).add_sample(cellvalue)
                                (objects[current_column]).add_location(plateloc)
                        if cellvalue == 'RNTC_NTC_A_1_1':                            
                            # checks if there is another assay title tucked under the end of this one
                            celllower = ('{}{}'.format(chr(64 + columncounter), (bottom_rows[bottomrowcounter])))
                            titlelower = ws[celllower].value
                            if titlelower != current_column and titlelower != None:
                                columntitles[columncounter - 3] = titlelower
                                if columntitles[columncounter - 3] not in objects:
                                    objects[titlelower] = Assay()
                    temprowcounter += 1
                try:
                    next_column = columntitles[columncounter-2]
                except IndexError:
                    # meaning were at the end of a platemap
                    if len(tucked_samples) != 0:
                        for sample in tucked_samples:
                            (objects[current_column]).add_sample(sample)
                        for location in tucked_locations:
                            (objects[current_column]).add_location(location)
                        tucked_samples.clear()
                        tucked_locations.clear()
                columncounter += 1
        if rowcounter > bottom_rows[bottomrowcounter]:
                bottomrowcounter += 1
        if rowcounter > bottom_rows[botrowlen - 1]:
            break  
    return objects


def edittemplate(ws, objects, gel, str1, gellen):
    # all of the output editing
    rowcounter = 0
    current_assay = 0
    # starts with an assay that belongs to the geltype you're looking for
    while objects[assaylist[current_assay]].geltype != gel:
        current_assay += 1
    samplenum = 0
    # default first control sample to AA
    control = 'AA'
    # later used to fill rest of gel with blank spaces when set to true
    blank = False
    # for index errors
    bust_samples = False
    # the shadow realm for paired assays to make sure they don't appear twice on the gel maps
    used_assays = []
    multichanneled = False
    totalsamples = 0    
    for row in ws.iter_rows():
        rowcounter += 1
        totalsamples = len(objects[assaylist[current_assay]].samples)
        # used for digested samples to appear twice
        turns = ""
        # when we get to multichanneling, need boolean to swap between assays
        using_paired_assay = False
        if (rowcounter-3)%(gellen+4) == 0:
            blank = False
            for x in range(gellen):
                # goes to next assay if used all samples from this assay
                if samplenum == totalsamples or bust_samples:
                    # if at the end of an assay, go to next assay
                    current_assay += 1
                    try:
                        while objects[assaylist[current_assay]].geltype != gel or assaylist[current_assay] in used_assays:
                            if current_assay == len(assaylist) +1:
                                return
                            current_assay += 1
                        totalsamples = len(objects[assaylist[current_assay]].samples)
                        if objects[assaylist[current_assay]].paired_assay != '':
                            # refering to the external reference information pulled earlier
                            multichanneled = True
                            paired_ass = objects[assaylist[current_assay]].paired_assay
                            used_assays.append(paired_ass)
                        else:
                            multichanneled = False  
                        if multichanneled:
                            # check for uneven sample counts between paired assays, suggesting a sample appears on only one of the assays
                            if totalsamples > len(objects[paired_ass].samples):
                                print('You will need to add the unique samples or NTC to {} manually'.format(assaylist[current_assay]))
                                totalsamples -= 1
                            if totalsamples < len(objects[paired_ass].samples):
                                print('You will need to add the unique samples or NTC to {} manually'.format(paired_ass))
                        samplenum = 0
                        teslen = ((1 + int(objects[assaylist[current_assay]].digested==True))*(totalsamples-1) + 1)
                        control = 'AA'
                        if current_assay == len(assaylist) +1:
                            return
                        if ((gellen-x) < teslen):
                            # if you can't fit the next assay on the rest of the gel, leave the rest empty.    
                            blank = True
                        bust_samples = False
                    except IndexError:
                        # leaves a note on the gel to grab any anomalies manually
                        ws.cell(row=(rowcounter+x), column=6).value = 'Last assay covered : {}. Make sure all assays were output'.format(assaylist[current_assay-1])
                        ws.cell(row=(rowcounter+x+1), column=6).value = 'Add unique samples and samples not at 20 ng'
                        return
                if blank == True:
                    # sets all values to none for the line
                    output_line(worksheet=ws, rows=rowcounter+x, blank=True)
                elif multichanneled:
                    # alternating between the paired assays
                    if not using_paired_assay:
                        assay_in_use = assaylist[current_assay]
                    else:
                        assay_in_use = paired_ass
                    try:
                        str1 = objects[assay_in_use].samples[samplenum]
                        if samplenum > 0:
                            if str1[0] == 'R':
                                control = 'AB'
                            else:
                                control = ''
                        if objects[assay_in_use].samples[samplenum] == "RNTC_NTC_A_1_1":
                            control = 'NTC'
                        output_line(
                            worksheet=ws,
                            rows=rowcounter+x,
                            assay=assay_in_use,
                            digested=turns,
                            control=control,
                            samplenum=samplenum,
                            objects=objects
                        )
                        if using_paired_assay:
                            # only move to next sample once paired assay has also had sample used
                            samplenum += 1
                        # move to the other assay
                        using_paired_assay = not using_paired_assay
                    except IndexError:
                        str2 = 'unique samples in {} or {} need to be manually added to the gel map'.format(assay_in_use, paired_ass)
                        ws.cell(row=(rowcounter+x-1), column=10).value = str2
                        bust_samples = True
                        samplenum += 1
                else:
                    # not multichanneled, could be digested but not necessary
                    if samplenum > 0:
                        str1 = objects[assaylist[current_assay]].samples[samplenum]
                        if str1[0] == 'R':
                            # check first letter of sample for 'R'
                            control = 'AB'
                            # remember first sample with R is already default to AA control. This is assuming it's the next control
                        else:
                            # assumed to be regular sample if no research number
                            control = ''
                        if objects[assaylist[current_assay]].samples[samplenum] == "RNTC_NTC_A_1_1":
                            control = 'NTC'
                    output_line(
                        worksheet=ws,
                        rows=rowcounter+x,
                        assay=assaylist[current_assay],
                        digested=turns,
                        control=control,
                        samplenum=samplenum,
                        objects=objects
                    )
                    if objects[assaylist[current_assay]].digested == True and turns == "" and objects[assaylist[current_assay]].samples[samplenum] != "RNTC_NTC_A_1_1":
                        # alternating indicator for digests
                        turns = "X"
                    elif objects[assaylist[current_assay]].samples[samplenum] == "RNTC_NTC_A_1_1":
                        samplenum += 1
                    else:
                        samplenum += 1
                        turns = ""                        


def output_line(worksheet, rows, sample=None, disease=None, assay=None, amount=None, digested=None, control=None, location=None, AA=None, AB=None, BB=None, samplenum=None, blank=False, objects=None):
    # puts all information into one line on the sheet
    if not blank:
        disease = objects[assay].disease
        sample=objects[assay].samples[samplenum]
        AA = objects[assay].A_allele
        AB = objects[assay].AB_allele
        BB = objects[assay].B_allele
        location = objects[assay].locations[samplenum]
        amount = 20
    worksheet.cell(row=rows, column=2).value = sample
    worksheet.cell(row=rows, column=3).value = disease
    worksheet.cell(row=rows, column=4).value = assay
    worksheet.cell(row=rows, column=5).value = amount
    worksheet.cell(row=rows, column=6).value = digested
    worksheet.cell(row=rows, column=7).value = control
    worksheet.cell(row=rows, column=9).value = location
    worksheet.cell(row=rows, column=11).value = AA
    worksheet.cell(row=rows, column=12).value = AB
    worksheet.cell(row=rows, column=13).value = BB


def output(path, objects, original_directory):
    os.chdir(original_directory)
    # each map that wants to be edited. one for each gel type
    templates = [
        'Sec_2%_E-GEL_Map.xlsx',
        'Sec_4%_E-GEL_Map.xlsx',
        'Sec_Agarose_Gel_Map.xlsx'
    ]
    geltypes = [
        "2% E-Gel",
        '4% E-Gel',
        'Agarose Gel'
    ]
    gel_lengths = {
        "2% E-Gel":12,
        '4% E-Gel':10,
        'Agarose Gel':12
    }
    mapcount = 0
    for template in templates:
        wb = load_workbook(template)
        ws = wb.active
        geltype = geltypes[mapcount]
        gel_length = gel_lengths[geltype]
        output_name = str(datetime.date.today()) + " " + str(template)
        edittemplate(ws, objects, geltype, output_name, gel_length)
        wb.save(output_name)
        move(output_name, path)
        mapcount += 1
    enda = input("Press enter to end this program")


if __name__ == "__main__":
    main()
